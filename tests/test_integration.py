"""
Integration tests for the forensic message analyzer.
"""

import pytest
import base64
import copy
import pandas as pd
from pathlib import Path
from PIL import Image
import tempfile
import json
import openpyxl
from datetime import datetime

from src.config import Config
from src.forensic_utils import ForensicRecorder, ForensicIntegrity
from src.extractors.data_extractor import DataExtractor
from src.analyzers.threat_analyzer import ThreatAnalyzer
from src.analyzers.sentiment_analyzer import SentimentAnalyzer
from src.analyzers.behavioral_analyzer import BehavioralAnalyzer
from src.analyzers.yaml_pattern_analyzer import YamlPatternAnalyzer
from src.analyzers.communication_metrics import CommunicationMetricsAnalyzer
from src.analyzers.attachment_processor import AttachmentProcessor
from src.review.manual_review_manager import ManualReviewManager
from src.reporters.forensic_reporter import ForensicReporter
from src.reporters.excel_reporter import ExcelReporter
from src.reporters.html_reporter import HtmlReporter
from src.reporters.json_reporter import JSONReporter
from src.main import ForensicAnalyzer


def _create_test_image(path: Path, width=10, height=10, color=(255, 0, 0)):
    """Create a small PNG test image."""
    img = Image.new('RGB', (width, height), color=color)
    img.save(path, format='PNG')
    return path


class TestSystemIntegration:
    """Test system integration and workflow."""
    
    def test_extraction_to_analysis_pipeline(self, tmp_path):
        """Test data flow from extraction to analysis."""
        recorder = ForensicRecorder(tmp_path)

        # Create test message data with clear threat and non-threat content
        test_messages = pd.DataFrame({
            'content': [
                'Hello, how are you?',
                'I will hurt you',
                'Normal conversation',
                'I am going to kill you'
            ],
            'sender': ['user1', 'user2', 'user1', 'user2'],
            'timestamp': pd.date_range(start='2024-01-01', periods=4, freq='h'),
            'source': ['iMessage'] * 4
        })

        # Run threat analysis
        threat_analyzer = ThreatAnalyzer(recorder)
        threat_results = threat_analyzer.detect_threats(test_messages)

        assert isinstance(threat_results, pd.DataFrame)
        assert len(threat_results) == len(test_messages)
        assert 'threat_detected' in threat_results.columns

        # Verify specific threat detection
        assert threat_results.loc[1, 'threat_detected'] == True, \
            "'I will hurt you' should be detected as a threat"
        assert threat_results.loc[3, 'threat_detected'] == True, \
            "'I am going to kill you' should be detected as a threat"
        assert threat_results.loc[0, 'threat_detected'] == False, \
            "'Hello, how are you?' should not be a threat"

        # Run sentiment analysis
        sentiment_analyzer = SentimentAnalyzer(recorder)
        sentiment_results = sentiment_analyzer.analyze_sentiment(test_messages)

        assert isinstance(sentiment_results, pd.DataFrame)
        assert 'sentiment_polarity' in sentiment_results.columns
    
    def test_analysis_to_review_pipeline(self, tmp_path):
        """Test data flow from analysis to manual review."""
        recorder = ForensicRecorder(tmp_path)

        # Create test data with clear threats
        test_messages = pd.DataFrame({
            'message_id': ['msg_001', 'msg_002', 'msg_003'],
            'content': [
                'I will hurt you',
                'Normal message',
                'I am going to kill you'
            ],
            'sender': ['user1', 'user2', 'user1'],
            'timestamp': pd.date_range(start='2024-01-01', periods=3, freq='h')
        })

        # Analyze threats
        threat_analyzer = ThreatAnalyzer(recorder)
        threat_results = threat_analyzer.detect_threats(test_messages)

        # Verify threats are detected before testing review pipeline
        assert threat_results['threat_detected'].any(), \
            "At least one threat should be detected for this test to be meaningful"

        # Create review manager
        review_dir = tmp_path / "reviews"
        review_manager = ManualReviewManager(review_dir=review_dir, forensic_recorder=recorder)

        # Add reviews for threats
        for idx, row in threat_results.iterrows():
            if row.get('threat_detected', False):
                review_manager.add_review(
                    test_messages.loc[idx, 'message_id'],
                    'threat',
                    'relevant',
                    f"Threat detected in message"
                )

        # Verify reviews were created for threat messages
        reviews = review_manager.get_reviews_by_decision('relevant')
        assert len(reviews) >= 2, \
            "Both 'I will hurt you' and 'I am going to kill you' should produce reviews"
        reviewed_ids = {r['item_id'] for r in reviews}
        assert 'msg_001' in reviewed_ids, "'I will hurt you' (msg_001) should be reviewed"
        assert 'msg_003' in reviewed_ids, "'I am going to kill you' (msg_003) should be reviewed"
        
    def test_review_manager(self, tmp_path):
        """Test manual review manager functionality."""
        recorder = ForensicRecorder(tmp_path)
        review_dir = tmp_path / "reviews"
        manager = ManualReviewManager(review_dir=review_dir, forensic_recorder=recorder)
        
        # Add multiple reviews
        manager.add_review('item1', 'threat', 'relevant', 'Contains threat')
        manager.add_review('item2', 'pattern', 'not_relevant', 'False positive')
        manager.add_review('item3', 'behavioral', 'uncertain', 'Needs more context')
        
        # Test retrieval by decision
        relevant = manager.get_reviews_by_decision('relevant')
        assert len(relevant) == 1
        assert relevant[0]['item_id'] == 'item1'
        
        not_relevant = manager.get_reviews_by_decision('not_relevant')
        assert len(not_relevant) == 1
        assert not_relevant[0]['item_id'] == 'item2'
        
        uncertain = manager.get_reviews_by_decision('uncertain')
        assert len(uncertain) == 1
        assert uncertain[0]['item_id'] == 'item3'
        
        # Test retrieval by type
        threat_reviews = manager.get_reviews_by_type('threat')
        assert len(threat_reviews) == 1
        assert threat_reviews[0]['decision'] == 'relevant'
        
        # Test summary
        summary = manager.get_review_summary()
        assert summary['total_reviews'] == 3
        assert summary['decisions']['relevant'] == 1
        assert summary['decisions']['not_relevant'] == 1
        assert summary['decisions']['uncertain'] == 1
        
    def test_full_workflow_integration(self, tmp_path):
        """
        Full end-to-end test: synthetic iMessage-style messages through
        extraction → analysis → review (with random mixed decisions) →
        filtering → report generation.

        Includes: image attachments, tapback reactions, emoji content,
        SOS/edited/retracted/downgraded flags, thread replies,
        HTML-hostile content, and missing-attachment fallback.

        Verifies that all report formats are produced with realistic data,
        that review filtering actually removes rejected findings, and that
        report contents correctly reflect special message types.
        """
        # ---------------------------------------------------------------
        # 0. Load production config — use the same env vars as a real run
        # ---------------------------------------------------------------
        config = Config()
        p1 = config.person1_name  # e.g. "David Snyder"
        p2 = config.person2_name  # e.g. "Marcia Snyder"

        # ---------------------------------------------------------------
        # 0b. Create test image fixtures for attachment testing
        # ---------------------------------------------------------------
        attachments_dir = tmp_path / "attachments"
        attachments_dir.mkdir()
        test_image_path = attachments_dir / "IMG_4521.png"
        _create_test_image(test_image_path, width=20, height=15, color=(255, 0, 0))
        test_jpeg_path = attachments_dir / "photo_evidence.jpg"
        _create_test_image(test_jpeg_path, width=30, height=20, color=(0, 0, 255))
        bruise_photo_path = attachments_dir / "bruise_photo.png"
        _create_test_image(bruise_photo_path, width=40, height=30, color=(128, 0, 128))
        screenshot_path = attachments_dir / "text_screenshot.png"
        _create_test_image(screenshot_path, width=60, height=100, color=(200, 200, 200))
        door_damage_path = attachments_dir / "door_damage.jpg"
        _create_test_image(door_damage_path, width=50, height=40, color=(139, 69, 19))

        # ---------------------------------------------------------------
        # 1. Build synthetic iMessage-style messages using the production
        #    person names so reports render exactly as they will in the
        #    real run. Covers: threats, normal conversation, images,
        #    tapbacks, emojis, and special message types.
        # ---------------------------------------------------------------
        messages = [
            # --- Original text messages (threats + normal) ---
            {
                'message_id': 'imsg_001',
                'guid': 'guid_001',
                'content': 'Hey, can you pick up the kids from school today?',
                'sender': p1,
                'recipient': p2,
                'timestamp': '2024-06-15T08:30:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
            },
            {
                'message_id': 'imsg_002',
                'guid': 'guid_002',
                'content': 'I told you I will hurt you if you keep ignoring my calls',
                'sender': p2,
                'recipient': p1,
                'timestamp': '2024-06-15T09:15:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
            },
            {
                'message_id': 'imsg_003',
                'guid': 'guid_003',
                'content': 'Please stop threatening me. I am documenting everything.',
                'sender': p1,
                'recipient': p2,
                'timestamp': '2024-06-15T09:20:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
            },
            {
                'message_id': 'imsg_004',
                'guid': 'guid_004',
                'content': 'You are worthless and nobody will believe you',
                'sender': p2,
                'recipient': p1,
                'timestamp': '2024-06-15T09:45:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
            },
            {
                'message_id': 'imsg_005',
                'guid': 'guid_005',
                'content': 'I will take the kids and you will never see them again',
                'sender': p2,
                'recipient': p1,
                'timestamp': '2024-06-15T10:00:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
            },
            {
                'message_id': 'imsg_006',
                'guid': 'guid_006',
                'content': 'Can we please talk about this calmly?',
                'sender': p1,
                'recipient': p2,
                'timestamp': '2024-06-15T10:30:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
            },
            {
                'message_id': 'imsg_007',
                'guid': 'guid_007',
                'content': 'I am going to destroy your car if you leave',
                'sender': p2,
                'recipient': p1,
                'timestamp': '2024-06-15T11:00:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
            },
            {
                'message_id': 'imsg_008',
                'guid': 'guid_008',
                'content': 'Sure, I will pick them up at 3pm',
                'sender': p1,
                'recipient': p2,
                'timestamp': '2024-06-15T14:00:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
            },
            {
                'message_id': 'imsg_009',
                'guid': 'guid_009',
                'content': 'You are crazy and insane for thinking that',
                'sender': p2,
                'recipient': p1,
                'timestamp': '2024-06-15T15:30:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
            },
            {
                'message_id': 'imsg_010',
                'guid': 'guid_010',
                'content': 'I hope we can work something out for the kids sake',
                'sender': p1,
                'recipient': p2,
                'timestamp': '2024-06-15T16:00:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
            },
            # --- Image attachment: photo of car damage ---
            {
                'message_id': 'imsg_011',
                'guid': 'guid_011',
                'content': 'Look at what he did to my car',
                'sender': p1,
                'recipient': p2,
                'timestamp': '2024-06-15T16:30:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
                'attachment': str(test_image_path),
                'attachment_name': 'IMG_4521.png',
                'attachments': [{
                    'path': str(test_image_path),
                    'name': 'IMG_4521.png',
                    'mime_type': 'image/png',
                    'size_bytes': test_image_path.stat().st_size,
                }],
            },
            # --- Tapback: Love reaction on the threat (imsg_002) ---
            {
                'message_id': 'imsg_012',
                'guid': 'guid_012',
                'content': 'Loved "I told you I will hurt you if you keep ignoring my calls"',
                'sender': p2,
                'recipient': p1,
                'timestamp': '2024-06-15T09:16:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
                'is_tapback': True,
                'associated_message_guid': 'p:0/guid_002',
                'associated_message_type': 2000,
                'associated_message_emoji': '\u2764\ufe0f',
            },
            # --- Tapback: Question reaction on custody threat (imsg_005) ---
            {
                'message_id': 'imsg_013',
                'guid': 'guid_013',
                'content': 'Questioned "I will take the kids and you will never see them again"',
                'sender': p1,
                'recipient': p2,
                'timestamp': '2024-06-15T10:01:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
                'is_tapback': True,
                'associated_message_guid': 'p:0/guid_005',
                'associated_message_type': 2005,
                'associated_message_emoji': '\u2753',
            },
            # --- Emoji-only message (angry face) ---
            {
                'message_id': 'imsg_014',
                'guid': 'guid_014',
                'content': '\U0001f621\U0001f92c\U0001f480',
                'sender': p2,
                'recipient': p1,
                'timestamp': '2024-06-15T11:30:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
            },
            # --- Threat with emojis ---
            {
                'message_id': 'imsg_015',
                'guid': 'guid_015',
                'content': 'I will kill you \U0001f608\U0001f52a',
                'sender': p2,
                'recipient': p1,
                'timestamp': '2024-06-15T12:00:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
            },
            # --- SOS emergency message ---
            {
                'message_id': 'imsg_016',
                'guid': 'guid_016',
                'content': 'Emergency SOS activated',
                'sender': p1,
                'recipient': p2,
                'timestamp': '2024-06-15T12:30:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
                'is_sos': True,
            },
            # --- Edited message ---
            {
                'message_id': 'imsg_017',
                'guid': 'guid_017',
                'content': 'I meant to say I am sorry',
                'sender': p2,
                'recipient': p1,
                'timestamp': '2024-06-15T13:00:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
                'date_edited': '2024-06-15T13:01:00+00:00',
            },
            # --- Retracted/unsent message ---
            {
                'message_id': 'imsg_018',
                'guid': 'guid_018',
                'content': 'This message was unsent',
                'sender': p2,
                'recipient': p1,
                'timestamp': '2024-06-15T13:30:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
                'date_retracted': '2024-06-15T13:31:00+00:00',
            },
            # --- Downgraded to SMS ---
            {
                'message_id': 'imsg_019',
                'guid': 'guid_019',
                'content': 'Text fell back to SMS',
                'sender': p1,
                'recipient': p2,
                'timestamp': '2024-06-15T13:45:00+00:00',
                'source': 'iMessage',
                'service': 'SMS',
                'was_downgraded': True,
            },
            # --- Message with reactions from others ---
            {
                'message_id': 'imsg_020',
                'guid': 'guid_020',
                'content': 'I called the police',
                'sender': p1,
                'recipient': p2,
                'timestamp': '2024-06-15T14:30:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
                'reactions': [
                    {'type': '\U0001f44d', 'sender': p2,
                     'timestamp': '2024-06-15T14:31:00+00:00'},
                    {'type': '\u2764\ufe0f', 'sender': p2,
                     'timestamp': '2024-06-15T14:32:00+00:00'},
                ],
            },
            # --- HTML-hostile content (XSS test + threat) ---
            {
                'message_id': 'imsg_021',
                'guid': 'guid_021',
                'content': '<script>alert("xss")</script> I will stalk you',
                'sender': p2,
                'recipient': p1,
                'timestamp': '2024-06-15T15:00:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
            },
            # --- Thread reply ---
            {
                'message_id': 'imsg_022',
                'guid': 'guid_022',
                'content': 'Replying to your earlier message about the kids',
                'sender': p1,
                'recipient': p2,
                'timestamp': '2024-06-15T15:15:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
                'thread_originator_guid': 'guid_001',
                'reply_to_guid': 'guid_005',
            },
            # --- Missing attachment (no file on disk, just name) ---
            {
                'message_id': 'imsg_023',
                'guid': 'guid_023',
                'content': 'Here is a photo',
                'sender': p2,
                'recipient': p1,
                'timestamp': '2024-06-15T15:45:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
                'attachment_name': 'missing_photo.jpg',
            },
            # --- Second image attachment (JPEG) ---
            {
                'message_id': 'imsg_024',
                'guid': 'guid_024',
                'content': 'Screenshot of his texts to my sister',
                'sender': p1,
                'recipient': p2,
                'timestamp': '2024-06-15T16:45:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
                'attachment': str(test_jpeg_path),
                'attachment_name': 'photo_evidence.jpg',
                'attachments': [{
                    'path': str(test_jpeg_path),
                    'name': 'photo_evidence.jpg',
                    'mime_type': 'image/jpeg',
                    'size_bytes': test_jpeg_path.stat().st_size,
                }],
            },

            # ============================================================
            # Day 2 — June 16 — escalation continues with more images
            # ============================================================
            {
                'message_id': 'imsg_025',
                'guid': 'guid_025',
                'content': 'I barely slept last night. Are you going to apologize?',
                'sender': p1,
                'recipient': p2,
                'timestamp': '2024-06-16T07:00:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
            },
            {
                'message_id': 'imsg_026',
                'guid': 'guid_026',
                'content': 'You deserve everything that is coming to you',
                'sender': p2,
                'recipient': p1,
                'timestamp': '2024-06-16T07:15:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
            },
            # --- Bruise photo with image attachment ---
            {
                'message_id': 'imsg_027',
                'guid': 'guid_027',
                'content': 'Look at this bruise from yesterday',
                'sender': p1,
                'recipient': p2,
                'timestamp': '2024-06-16T07:30:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
                'attachment': str(bruise_photo_path),
                'attachment_name': 'bruise_photo.png',
                'attachments': [{
                    'path': str(bruise_photo_path),
                    'name': 'bruise_photo.png',
                    'mime_type': 'image/png',
                    'size_bytes': bruise_photo_path.stat().st_size,
                }],
            },
            {
                'message_id': 'imsg_028',
                'guid': 'guid_028',
                'content': 'That was your own fault and you know it',
                'sender': p2,
                'recipient': p1,
                'timestamp': '2024-06-16T07:45:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
            },
            {
                'message_id': 'imsg_029',
                'guid': 'guid_029',
                'content': 'I am taking the kids to my mothers house today',
                'sender': p1,
                'recipient': p2,
                'timestamp': '2024-06-16T08:00:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
            },
            {
                'message_id': 'imsg_030',
                'guid': 'guid_030',
                'content': 'If you take those kids I will burn the house down',
                'sender': p2,
                'recipient': p1,
                'timestamp': '2024-06-16T08:05:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
            },
            # --- Tapback: Dislike on the arson threat (imsg_030) ---
            {
                'message_id': 'imsg_031',
                'guid': 'guid_031',
                'content': 'Disliked "If you take those kids I will burn the house down"',
                'sender': p1,
                'recipient': p2,
                'timestamp': '2024-06-16T08:06:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
                'is_tapback': True,
                'associated_message_guid': 'p:0/guid_030',
                'associated_message_type': 3001,
                'associated_message_emoji': '\U0001f44e',
            },
            {
                'message_id': 'imsg_032',
                'guid': 'guid_032',
                'content': 'My attorney said you need to stop contacting me directly',
                'sender': p1,
                'recipient': p2,
                'timestamp': '2024-06-16T09:00:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
            },
            {
                'message_id': 'imsg_033',
                'guid': 'guid_033',
                'content': 'Your attorney can go to hell',
                'sender': p2,
                'recipient': p1,
                'timestamp': '2024-06-16T09:10:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
            },
            # --- Screenshot of prior text conversation, with image ---
            {
                'message_id': 'imsg_034',
                'guid': 'guid_034',
                'content': 'Here is the screenshot of what you sent my mom',
                'sender': p1,
                'recipient': p2,
                'timestamp': '2024-06-16T09:30:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
                'attachment': str(screenshot_path),
                'attachment_name': 'text_screenshot.png',
                'attachments': [{
                    'path': str(screenshot_path),
                    'name': 'text_screenshot.png',
                    'mime_type': 'image/png',
                    'size_bytes': screenshot_path.stat().st_size,
                }],
            },
            {
                'message_id': 'imsg_035',
                'guid': 'guid_035',
                'content': 'I never sent that, you are making things up again',
                'sender': p2,
                'recipient': p1,
                'timestamp': '2024-06-16T09:45:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
            },
            {
                'message_id': 'imsg_036',
                'guid': 'guid_036',
                'content': 'I have the receipts. Stop gaslighting me.',
                'sender': p1,
                'recipient': p2,
                'timestamp': '2024-06-16T10:00:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
            },
            {
                'message_id': 'imsg_037',
                'guid': 'guid_037',
                'content': 'I will make sure you regret this',
                'sender': p2,
                'recipient': p1,
                'timestamp': '2024-06-16T10:30:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
            },
            # --- Door damage photo ---
            {
                'message_id': 'imsg_038',
                'guid': 'guid_038',
                'content': 'He kicked in the bedroom door this morning',
                'sender': p1,
                'recipient': p2,
                'timestamp': '2024-06-16T11:00:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
                'attachment': str(door_damage_path),
                'attachment_name': 'door_damage.jpg',
                'attachments': [{
                    'path': str(door_damage_path),
                    'name': 'door_damage.jpg',
                    'mime_type': 'image/jpeg',
                    'size_bytes': door_damage_path.stat().st_size,
                }],
            },
            {
                'message_id': 'imsg_039',
                'guid': 'guid_039',
                'content': 'That door was already broken',
                'sender': p2,
                'recipient': p1,
                'timestamp': '2024-06-16T11:15:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
            },
            {
                'message_id': 'imsg_040',
                'guid': 'guid_040',
                'content': 'The kids are scared and asking why daddy is so angry',
                'sender': p1,
                'recipient': p2,
                'timestamp': '2024-06-16T12:00:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
            },
            {
                'message_id': 'imsg_041',
                'guid': 'guid_041',
                'content': 'Tell them to mind their own business',
                'sender': p2,
                'recipient': p1,
                'timestamp': '2024-06-16T12:15:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
            },
            {
                'message_id': 'imsg_042',
                'guid': 'guid_042',
                'content': 'They are children! They should not have to live like this',
                'sender': p1,
                'recipient': p2,
                'timestamp': '2024-06-16T12:30:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
            },
            {
                'message_id': 'imsg_043',
                'guid': 'guid_043',
                'content': 'I am going to find you wherever you go \U0001f440',
                'sender': p2,
                'recipient': p1,
                'timestamp': '2024-06-16T13:00:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
            },
            # --- Another edited message (day 2) ---
            {
                'message_id': 'imsg_044',
                'guid': 'guid_044',
                'content': 'I did not mean that last message',
                'sender': p2,
                'recipient': p1,
                'timestamp': '2024-06-16T13:05:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
                'date_edited': '2024-06-16T13:06:00+00:00',
            },
            {
                'message_id': 'imsg_045',
                'guid': 'guid_045',
                'content': 'I filed the protective order today',
                'sender': p1,
                'recipient': p2,
                'timestamp': '2024-06-16T15:00:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
            },
            {
                'message_id': 'imsg_046',
                'guid': 'guid_046',
                'content': 'A piece of paper wont protect you',
                'sender': p2,
                'recipient': p1,
                'timestamp': '2024-06-16T15:30:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
            },
            # --- Tapback: Emphasis on protective order message ---
            {
                'message_id': 'imsg_047',
                'guid': 'guid_047',
                'content': 'Emphasized "I filed the protective order today"',
                'sender': p1,
                'recipient': p2,
                'timestamp': '2024-06-16T15:01:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
                'is_tapback': True,
                'associated_message_guid': 'p:0/guid_045',
                'associated_message_type': 2004,
                'associated_message_emoji': '!!',
            },
            {
                'message_id': 'imsg_048',
                'guid': 'guid_048',
                'content': 'Please just leave us alone. The kids need peace.',
                'sender': p1,
                'recipient': p2,
                'timestamp': '2024-06-16T16:00:00+00:00',
                'source': 'iMessage',
                'service': 'iMessage',
            },
        ]

        # JSON round-trip to match real pipeline behaviour
        extracted_data = json.loads(json.dumps({
            'messages': messages,
            'screenshots': [],
            'combined': messages,
            'third_party_contacts': [],
        }, default=str))

        # ---------------------------------------------------------------
        # 2. Preserve attachments (FRE 1002 — mirrors production pipeline)
        # ---------------------------------------------------------------
        temp_dir = tmp_path / "output"
        temp_dir.mkdir()
        forensic = ForensicRecorder(temp_dir)

        integrity = ForensicIntegrity(forensic)
        att_dest_dir = temp_dir / "attachments"
        preserved = {}          # {original_path_str: preserved_path}
        original_hashes = {}    # {original_path_str: sha256_hash}

        for msg in extracted_data['messages']:
            att_path_str = msg.get('attachment')
            if att_path_str:
                att_path = Path(att_path_str)
                if att_path_str in preserved:
                    msg['attachment'] = str(preserved[att_path_str])
                elif att_path.is_file():
                    original_hashes[att_path_str] = forensic.compute_hash(att_path)
                    copy_path = integrity.create_working_copy(att_path, att_dest_dir)
                    assert copy_path is not None, f"Failed to preserve {att_path.name}"
                    preserved[att_path_str] = copy_path
                    msg['attachment'] = str(copy_path)

            for att in msg.get('attachments', []):
                att_list_path_str = att.get('path')
                if not att_list_path_str:
                    continue
                att_list_path = Path(att_list_path_str)
                if att_list_path_str in preserved:
                    att['path'] = str(preserved[att_list_path_str])
                elif att_list_path.is_file():
                    if att_list_path_str not in original_hashes:
                        original_hashes[att_list_path_str] = forensic.compute_hash(att_list_path)
                    copy_path = integrity.create_working_copy(att_list_path, att_dest_dir)
                    assert copy_path is not None, f"Failed to preserve {att_list_path.name}"
                    preserved[att_list_path_str] = copy_path
                    att['path'] = str(copy_path)

        assert len(preserved) == 5, (
            f"Expected 5 unique attachment files preserved, got {len(preserved)}: "
            f"{[Path(str(p)).name for p in preserved.values()]}"
        )
        assert att_dest_dir.exists(), "Attachments output directory should exist"

        # ---------------------------------------------------------------
        # 3. Run all analysis phases
        # ---------------------------------------------------------------
        df = pd.DataFrame(messages)

        ta = ThreatAnalyzer(forensic)
        threat_df = ta.detect_threats(df)
        threat_summary = ta.generate_threat_summary(threat_df)

        sa = SentimentAnalyzer(forensic)
        sentiment_df = sa.analyze_sentiment(df)

        pa = YamlPatternAnalyzer(forensic)
        pattern_df = pa.analyze_patterns(df)

        ba = BehavioralAnalyzer(forensic)
        behavioral = ba.analyze_patterns(df)

        cm = CommunicationMetricsAnalyzer(forensic_recorder=forensic)
        metrics = cm.analyze_messages(messages)

        # Verify threat detection found the right messages
        threat_count = int(threat_df['threat_detected'].sum())
        assert threat_count >= 8, (
            f"Expected at least 8 threats from 48 synthetic messages, "
            f"got {threat_count}"
        )

        # Verify emoji-adjacent threat is detected
        emoji_threat_row = threat_df[threat_df['content'] == 'I will kill you \U0001f608\U0001f52a']
        assert len(emoji_threat_row) == 1, "Emoji threat message not found in DataFrame"
        assert emoji_threat_row.iloc[0]['threat_detected'] == True, (
            "'I will kill you [emojis]' should be detected as a threat"
        )

        # Verify HTML-hostile content threat is detected
        xss_threat_row = threat_df[
            threat_df['content'].str.contains('stalk', na=False)
        ]
        assert len(xss_threat_row) >= 1, "XSS+threat message not found"
        assert xss_threat_row.iloc[0]['threat_detected'] == True, (
            "'<script>...</script> I will stalk you' should be detected"
        )

        # Stub AI analysis with realistic structure including
        # threat details that the review filter can act on
        ai_analysis = {
            'generated_at': datetime.now().isoformat(),
            'total_messages': len(messages),
            'ai_model': 'test-stub',
            'conversation_summary': (
                f'Conversation between {p1} and {p2} shows escalating '
                f'conflict with multiple threats of physical harm and custody '
                f'interference by {p2}. {p1} documents evidence including '
                f'photos of property damage. An emergency SOS was activated.'
            ),
            'sentiment_analysis': {
                'scores': [{'batch': 1, 'avg_polarity': -0.35}],
                'overall': 'negative',
                'shifts': [{'from': 'neutral', 'to': 'hostile',
                            'at': '2024-06-15T09:15:00'}],
            },
            'threat_assessment': {
                'found': True,
                'details': [
                    {
                        'type': 'physical_threat',
                        'severity': 'high',
                        'quote': 'I will hurt you if you keep ignoring my calls',
                        'context': 'Direct threat of physical harm',
                    },
                    {
                        'type': 'custody_interference',
                        'severity': 'high',
                        'quote': 'I will take the kids and you will never see them',
                        'context': 'Threat to remove children',
                    },
                    {
                        'type': 'property_destruction',
                        'severity': 'medium',
                        'quote': 'I am going to destroy your car',
                        'context': 'Threat to damage property',
                    },
                    {
                        'type': 'stalking',
                        'severity': 'high',
                        'quote': 'I will stalk you',
                        'context': 'Stalking threat embedded in hostile message',
                    },
                    {
                        'type': 'arson_threat',
                        'severity': 'critical',
                        'quote': 'I will burn the house down',
                        'context': 'Threat to commit arson if children are removed',
                    },
                    {
                        'type': 'stalking',
                        'severity': 'high',
                        'quote': 'I am going to find you wherever you go',
                        'context': 'Continued stalking threat on day 2',
                    },
                    {
                        'type': 'intimidation',
                        'severity': 'high',
                        'quote': 'A piece of paper wont protect you',
                        'context': 'Dismissing protective order as ineffective',
                    },
                ],
            },
            'behavioral_patterns': {
                'intimidation': {'detected': True, 'frequency': 'repeated'},
                'emotional_abuse': {'detected': True, 'frequency': 'repeated'},
            },
            'key_topics': ['custody', 'threats', 'emotional abuse', 'property damage'],
            'risk_indicators': [
                {'type': 'threat', 'severity': 'high',
                 'detail': 'Multiple explicit threats of harm'},
                {'type': 'behavioral', 'severity': 'medium',
                 'detail': 'Pattern of emotional abuse and name-calling'},
                {'type': 'emergency', 'severity': 'critical',
                 'detail': 'Emergency SOS was triggered during conversation'},
            ],
            'notable_quotes': [
                {
                    'quote': 'I will hurt you if you keep ignoring my calls',
                    'context': 'Escalation after perceived rejection',
                    'significance': 'Direct threat',
                },
                {
                    'quote': 'You are worthless and nobody will believe you',
                    'context': 'Emotional manipulation',
                    'significance': 'Gaslighting attempt',
                },
                {
                    'quote': 'I will kill you \U0001f608\U0001f52a',
                    'context': 'Threat with menacing emojis',
                    'significance': 'Death threat with emoji emphasis',
                },
            ],
            'recommendations': [
                'Document all threatening communications',
                'Consider protective order based on repeated threats',
                'Preserve photo evidence of property damage',
            ],
            'processing_stats': {
                'batches_processed': 1,
                'tokens_used': 500,
                'input_tokens': 300,
                'output_tokens': 200,
                'api_calls': 1,
                'errors': [],
            },
        }

        analysis_results = {
            'threats': {
                'details': threat_df.to_dict('records'),
                'summary': threat_summary,
            },
            'sentiment': sentiment_df.to_dict('records'),
            'patterns': pattern_df.to_dict('records'),
            'metrics': metrics,
            'ai_analysis': ai_analysis,
        }

        # ---------------------------------------------------------------
        # 4. Build review items and apply mixed decisions
        # ---------------------------------------------------------------
        items_for_review = []
        threat_details = analysis_results['threats']['details']
        for idx, item in enumerate(threat_details):
            if item.get('threat_detected'):
                items_for_review.append({
                    'id': f"threat_{idx}",
                    'type': 'threat',
                    'content': item.get('content', ''),
                })

        ai_threats = ai_analysis['threat_assessment']['details']
        for i, detail in enumerate(ai_threats):
            items_for_review.append({
                'id': f"ai_threat_{i}",
                'type': 'ai_threat',
                'content': detail.get('quote', ''),
            })

        for i, nq in enumerate(ai_analysis.get('notable_quotes', [])):
            items_for_review.append({
                'id': f"ai_notable_{i}",
                'type': 'ai_notable',
                'content': nq.get('quote', ''),
            })

        assert len(items_for_review) >= 10, (
            f"Expected at least 10 review items, got {len(items_for_review)}"
        )

        # Cycle through relevant / not_relevant / uncertain to simulate
        # a realistic human review with mixed decisions
        review_dir = tmp_path / "reviews"
        manager = ManualReviewManager(review_dir=review_dir, forensic_recorder=forensic)
        decision_cycle = ['relevant', 'not_relevant', 'uncertain']
        for i, item in enumerate(items_for_review):
            decision = decision_cycle[i % 3]
            notes = {
                'relevant': 'Confirmed \u2014 include in report',
                'not_relevant': 'False positive \u2014 exclude',
                'uncertain': 'Needs attorney review',
            }[decision]
            manager.add_review(item['id'], item['type'], decision, notes=notes)

        review_results = {
            'total_reviewed': len(manager.reviews),
            'relevant': len(manager.get_reviews_by_decision('relevant')),
            'not_relevant': len(manager.get_reviews_by_decision('not_relevant')),
            'uncertain': len(manager.get_reviews_by_decision('uncertain')),
            'reviews': manager.reviews,
        }

        # Must have at least one of each decision type
        assert review_results['relevant'] >= 1
        assert review_results['not_relevant'] >= 1
        assert review_results['uncertain'] >= 1

        # ---------------------------------------------------------------
        # 5. Filter analysis by review decisions
        # ---------------------------------------------------------------
        config.output_dir = str(temp_dir)
        analyzer = ForensicAnalyzer(config)
        filtered_analysis = analyzer._filter_analysis_by_review(
            analysis_results, review_results
        )

        # Verify filtering actually removed rejected threats
        rejected_ids = {
            r['item_id'] for r in manager.get_reviews_by_decision('not_relevant')
        }
        for idx, item in enumerate(filtered_analysis['threats']['details']):
            if f"threat_{idx}" in rejected_ids:
                assert not item.get('threat_detected'), (
                    f"threat_{idx} was rejected in review but still shows "
                    f"threat_detected=True after filtering"
                )

        # ---------------------------------------------------------------
        # 6. Generate all report formats
        # ---------------------------------------------------------------
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Excel report
        excel_reporter = ExcelReporter(forensic, config=config)
        excel_path = temp_dir / f"report_{timestamp}.xlsx"
        excel_reporter.generate_report(
            extracted_data, filtered_analysis, review_results, excel_path
        )
        assert excel_path.exists(), "Excel report was not created"
        assert excel_path.stat().st_size > 0, "Excel report is empty"

        # HTML report (skip PDF to avoid WeasyPrint dependency in CI)
        html_reporter = HtmlReporter(forensic, config=config)
        html_base = temp_dir / f"report_{timestamp}"
        html_paths = html_reporter.generate_report(
            extracted_data, filtered_analysis, review_results,
            html_base, pdf=False
        )
        assert len(html_paths) >= 1, "HTML reporter produced no files"
        for fmt, path in html_paths.items():
            assert Path(path).exists(), f"HTML report file missing: {path}"

        # JSON report
        json_reporter = JSONReporter(forensic, config=config)
        json_path = temp_dir / f"report_{timestamp}.json"
        json_reporter.generate_report(
            extracted_data, filtered_analysis, review_results, json_path
        )
        assert json_path.exists(), "JSON report was not created"
        with open(json_path) as f:
            json_data = json.load(f)

        # ForensicReporter (Word + legal team summary)
        forensic_reporter = ForensicReporter(forensic, config=config)
        fr_reports = forensic_reporter.generate_comprehensive_report(
            extracted_data, filtered_analysis, review_results
        )
        assert len(fr_reports) >= 1, "ForensicReporter produced no files"
        for fmt, path in fr_reports.items():
            assert Path(path).exists(), f"Forensic report missing: {fmt} -> {path}"
            assert Path(path).stat().st_size > 0, f"Forensic report is empty: {fmt}"

        # Chain of custody
        chain_path = forensic.generate_chain_of_custody()
        assert Path(chain_path).exists(), "Chain of custody file not created"

        # ---------------------------------------------------------------
        # 7. Verify output directory has the expected files
        # ---------------------------------------------------------------
        output_files = [f for f in temp_dir.rglob("*") if f.is_file()]
        assert len(output_files) >= 5, (
            f"Expected at least 5 output files (Excel, HTML, JSON, Word, "
            f"chain of custody), got {len(output_files)}: "
            f"{[f.name for f in output_files]}"
        )

        # ---------------------------------------------------------------
        # 8. Verify HTML report content reflects special message types
        # ---------------------------------------------------------------
        html_path = html_paths.get('html')
        assert html_path, "No HTML path returned"
        html_content = Path(html_path).read_text(encoding='utf-8')

        # Image attachments: base64 data URIs should be embedded in original formats
        assert 'data:image/png;base64,' in html_content, (
            "HTML report should contain base64-embedded PNG image"
        )
        assert 'data:image/jpeg;base64,' in html_content, (
            "HTML report should contain base64-embedded JPEG image"
        )

        # Missing attachment: fallback text
        assert 'missing_photo.jpg' in html_content, (
            "HTML report should show fallback for missing attachment"
        )

        # Tapback styling: at least one "Tapback" label
        assert 'Tapback' in html_content, (
            "HTML report should contain 'Tapback' label for tapback messages"
        )

        # SOS flag
        assert 'SOS' in html_content, (
            "HTML report should contain SOS flag for emergency message"
        )

        # Edited flag
        assert 'Edited' in html_content, (
            "HTML report should contain 'Edited' flag for edited message"
        )

        # Retracted/Unsent flag
        assert 'Unsent' in html_content, (
            "HTML report should contain 'Unsent' flag for retracted message"
        )

        # Downgraded SMS flag
        assert 'SMS' in html_content, (
            "HTML report should contain 'SMS' flag for downgraded message"
        )

        # Reactions display
        assert p2 in html_content, (
            f"HTML report should show reaction sender ({p2})"
        )

        # HTML escaping: raw <script> tag must NOT appear
        assert '<script>' not in html_content, (
            "HTML report must escape <script> tags to prevent XSS"
        )
        # The escaped version should be present
        assert '&lt;script&gt;' in html_content or 'stalk' in html_content, (
            "HTML report should contain escaped version of hostile content"
        )

        # Emoji content should be preserved
        assert '\U0001f621' in html_content or '&#' in html_content, (
            "HTML report should contain emoji content"
        )

        # Thread reference
        assert 'guid_001' in html_content, (
            "HTML report should show thread originator GUID"
        )

        # ---------------------------------------------------------------
        # 9. Verify JSON report preserves special fields
        # ---------------------------------------------------------------
        json_messages = json_data.get('extraction', {}).get('messages', [])
        if not json_messages:
            json_messages = json_data.get('messages', [])

        # Find the tapback message in JSON
        tapback_msgs = [m for m in json_messages if m.get('is_tapback')]
        assert len(tapback_msgs) >= 1, (
            "JSON report should preserve tapback messages"
        )

        # Find the SOS message in JSON
        sos_msgs = [m for m in json_messages if m.get('is_sos')]
        assert len(sos_msgs) >= 1, (
            "JSON report should preserve SOS messages"
        )

        # Find messages with reactions
        reaction_msgs = [m for m in json_messages if m.get('reactions')]
        assert len(reaction_msgs) >= 1, (
            "JSON report should preserve reaction arrays"
        )

        # Find messages with attachments
        attachment_msgs = [m for m in json_messages if m.get('attachment')]
        assert len(attachment_msgs) >= 1, (
            "JSON report should preserve attachment paths"
        )

        # ---------------------------------------------------------------
        # 10. Verify Excel report has content and date fields
        # ---------------------------------------------------------------
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        sheet_names = wb.sheetnames
        assert len(sheet_names) >= 1, "Excel report has no sheets"
        # The Overview sheet should exist
        assert 'Overview' in sheet_names, (
            f"Excel report missing 'Overview' sheet, has: {sheet_names}"
        )

        # Read the Overview sheet and verify date fields are populated
        overview_ws = wb['Overview']
        overview_rows = list(overview_ws.iter_rows(min_row=2, values_only=True))
        overview_dict = {row[0]: row[1] for row in overview_rows if row[0]}

        # Date Range must show the actual message span, not "N/A"
        date_range_val = str(overview_dict.get('Date Range', ''))
        assert date_range_val != 'N/A', (
            f"Excel Date Range should be computed from messages, got: {date_range_val}"
        )
        assert '2024-06-15' in date_range_val, (
            f"Excel Date Range should include first day (2024-06-15), got: {date_range_val}"
        )
        assert '2024-06-16' in date_range_val, (
            f"Excel Date Range should include second day (2024-06-16), got: {date_range_val}"
        )

        # Report Generated must have an actual timestamp
        report_generated_val = str(overview_dict.get('Report Generated', ''))
        assert report_generated_val and report_generated_val != 'N/A', (
            f"Excel 'Report Generated' should have a timestamp, got: {report_generated_val}"
        )
        # Should contain a year (basic sanity check that it's a real date)
        assert '20' in report_generated_val, (
            f"Excel 'Report Generated' doesn't look like a date: {report_generated_val}"
        )

        # Total Messages should reflect the full 48-message dataset
        total_msg_val = overview_dict.get('Total Messages', 0)
        assert int(total_msg_val) >= 40, (
            f"Excel Total Messages should be >= 40, got: {total_msg_val}"
        )

        # Findings Summary sheet should exist with timestamped findings
        assert 'Findings Summary' in sheet_names, (
            f"Excel report missing 'Findings Summary' sheet, has: {sheet_names}"
        )
        findings_ws = wb['Findings Summary']
        findings_headers = [cell.value for cell in next(findings_ws.iter_rows(min_row=1, max_row=1))]
        assert 'Timestamp' in findings_headers, (
            f"Findings Summary should have 'Timestamp' column, has: {findings_headers}"
        )
        assert 'Section' in findings_headers, (
            f"Findings Summary should have 'Section' column, has: {findings_headers}"
        )
        # Should have at least one confirmed threat row
        findings_rows = list(findings_ws.iter_rows(min_row=2, values_only=True))
        findings_sections = [row[0] for row in findings_rows if row[0]]
        assert 'Confirmed Threat' in findings_sections, (
            f"Findings Summary should contain 'Confirmed Threat' rows, has sections: {findings_sections}"
        )

        # Timeline sheet should exist with chronological events
        assert 'Timeline' in sheet_names, (
            f"Excel report missing 'Timeline' sheet, has: {sheet_names}"
        )
        timeline_ws = wb['Timeline']
        timeline_headers = [cell.value for cell in next(timeline_ws.iter_rows(min_row=1, max_row=1))]
        assert 'Event Type' in timeline_headers, (
            f"Timeline should have 'Event Type' column, has: {timeline_headers}"
        )
        timeline_rows = list(timeline_ws.iter_rows(min_row=2, values_only=True))
        assert len(timeline_rows) > 0, "Timeline sheet should have at least one event"

        wb.close()

        # ---------------------------------------------------------------
        # 11. Verify HTML report date is populated
        # ---------------------------------------------------------------
        assert 'Report Date' in html_content, (
            "HTML report should contain 'Report Date' label"
        )
        # The report_date template variable should have rendered a real timestamp
        # (not an empty string). format_timestamp() produces "YYYY-MM-DD HH:MM:SS TZ".
        import re
        html_date_match = re.search(
            r'Report Date.*?(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})', html_content, re.DOTALL
        )
        assert html_date_match, (
            "HTML report 'Report Date' field should contain a rendered timestamp"
        )

        # ---------------------------------------------------------------
        # 12. Verify attachment preservation (FRE 1002)
        # ---------------------------------------------------------------
        # Preserved copies must exist with matching hashes
        for original_path_str, preserved_path in preserved.items():
            assert preserved_path.exists(), (
                f"Preserved attachment missing: {preserved_path}"
            )
            preserved_hash = forensic.compute_hash(preserved_path)
            assert preserved_hash == original_hashes[original_path_str], (
                f"Hash mismatch for preserved copy of {Path(original_path_str).name}: "
                f"original={original_hashes[original_path_str]}, "
                f"copy={preserved_hash}"
            )

        # Message dicts must reference preserved copies, not originals
        for msg in extracted_data['messages']:
            att_path = msg.get('attachment')
            if att_path:
                assert str(att_dest_dir) in att_path, (
                    f"Message attachment should reference preserved copy in "
                    f"{att_dest_dir}, got: {att_path}"
                )
            for att in msg.get('attachments', []):
                att_list_path = att.get('path')
                if att_list_path:
                    assert str(att_dest_dir) in att_list_path, (
                        f"Message attachments list should reference preserved copy, "
                        f"got: {att_list_path}"
                    )

        # Output directory should contain exactly 5 preserved files
        preserved_files = list(att_dest_dir.iterdir())
        assert len(preserved_files) == 5, (
            f"Expected 5 preserved files in attachments dir, got "
            f"{len(preserved_files)}: {[f.name for f in preserved_files]}"
        )

    # ------------------------------------------------------------------
    # Attachment processor test
    # ------------------------------------------------------------------

    def test_attachment_processor(self, tmp_path):
        """Test attachment processor with real image files."""
        forensic = ForensicRecorder(tmp_path)
        processor = AttachmentProcessor(forensic)

        # Create test image files
        att_dir = tmp_path / "test_attachments"
        att_dir.mkdir()

        png_path = att_dir / "test_image.png"
        _create_test_image(png_path, width=50, height=40, color=(0, 255, 0))

        jpeg_path = att_dir / "test_photo.jpg"
        img = Image.new('RGB', (100, 80), color=(0, 0, 255))
        img.save(jpeg_path, format='JPEG')

        # Create a small text file as non-image attachment
        txt_path = att_dir / "notes.txt"
        txt_path.write_text("Evidence notes for case review")

        # Process all attachments in directory
        results = processor.process_attachments(att_dir)
        assert len(results) == 3, f"Expected 3 attachments, got {len(results)}"

        # Verify each attachment has required fields
        for att in results:
            assert 'filename' in att
            assert 'file_hash' in att
            assert len(att['file_hash']) == 64, "SHA-256 hash should be 64 hex chars"
            assert 'size_bytes' in att
            assert att['size_bytes'] > 0
            assert 'mime_type' in att
            assert 'type' in att

        # Verify image metadata extraction
        png_results = [a for a in results if a['filename'] == 'test_image.png']
        assert len(png_results) == 1
        png_att = png_results[0]
        assert png_att['type'] == 'image'
        assert png_att['metadata'].get('width') == 50
        assert png_att['metadata'].get('height') == 40

        # Verify categorization
        types = {a['filename']: a['type'] for a in results}
        assert types['test_image.png'] == 'image'
        assert types['test_photo.jpg'] == 'image'
        assert types['notes.txt'] == 'text'

        # Test summary generation
        summary = processor.generate_attachment_summary(results)
        assert summary['total_attachments'] == 3
        assert summary['total_size_bytes'] > 0
        assert 'image' in summary['types']
        assert summary['types']['image'] == 2

        # Test single attachment processing
        single = processor.process_single_attachment(png_path)
        assert single['filename'] == 'test_image.png'
        assert single['file_hash'] == png_att['file_hash'], "Hash should be deterministic"

    def test_attachment_processor_missing_directory(self, tmp_path):
        """Test attachment processor with non-existent directory."""
        forensic = ForensicRecorder(tmp_path)
        processor = AttachmentProcessor(forensic)
        results = processor.process_attachments(tmp_path / "nonexistent")
        assert results == [], "Should return empty list for missing directory"

    # ------------------------------------------------------------------
    # Edge case tests for threat detection
    # ------------------------------------------------------------------

    def test_threat_detection_edge_cases(self, tmp_path):
        """Test threat detection with edge case content."""
        recorder = ForensicRecorder(tmp_path)
        analyzer = ThreatAnalyzer(recorder)

        test_data = pd.DataFrame({
            'content': [
                # Emoji-only (no threat words)
                '\U0001f621\U0001f92c\U0001f480',
                # Threat with emojis
                'I will kill you \U0001f608\U0001f52a',
                # HTML content with threat
                '<b>I will hurt you</b>',
                # Empty string
                '',
                # Unicode characters with threat
                'I\u2019ll destroy everything you own',
                # Normal emoji use
                'Love you \u2764\ufe0f',
                # URL with threat-like words (should not be a threat)
                'Check out https://www.killbill.com/movie',
            ],
            'sender': ['u1'] * 7,
            'timestamp': pd.date_range(start='2024-01-01', periods=7, freq='h'),
        })

        results = analyzer.detect_threats(test_data)
        assert len(results) == 7

        # Emoji-only: no threat words, should not be flagged
        assert results.loc[0, 'threat_detected'] == False, (
            "Emoji-only message should not be flagged as threat"
        )

        # Threat with emojis: 'kill' should still be detected
        assert results.loc[1, 'threat_detected'] == True, (
            "'I will kill you [emojis]' should be detected as threat"
        )

        # HTML content with threat: 'hurt' should be detected
        assert results.loc[2, 'threat_detected'] == True, (
            "'<b>I will hurt you</b>' should detect 'hurt' as threat"
        )

        # Empty string: should not crash or flag
        assert results.loc[3, 'threat_detected'] == False, (
            "Empty string should not be flagged as threat"
        )

    def test_threat_detection_nan_content(self, tmp_path):
        """Test threat detection handles NaN/None content gracefully."""
        recorder = ForensicRecorder(tmp_path)
        analyzer = ThreatAnalyzer(recorder)

        test_data = pd.DataFrame({
            'content': [None, float('nan'), 'Normal message', None, 'I will hurt you'],
            'sender': ['u1'] * 5,
            'timestamp': pd.date_range(start='2024-01-01', periods=5, freq='h'),
        })

        results = analyzer.detect_threats(test_data)
        assert len(results) == 5, "Should handle NaN/None without crashing"

        # None and NaN should not be flagged
        assert results.loc[0, 'threat_detected'] == False
        assert results.loc[1, 'threat_detected'] == False

        # Real threat should still be detected
        assert results.loc[4, 'threat_detected'] == True

    def test_sentiment_analysis_edge_cases(self, tmp_path):
        """Test sentiment analysis with emoji and edge case content."""
        recorder = ForensicRecorder(tmp_path)
        analyzer = SentimentAnalyzer(recorder)

        test_data = pd.DataFrame({
            'content': [
                '\U0001f621\U0001f92c',           # angry emojis
                '\u2764\ufe0f\U0001f60d',          # love emojis
                '',                                 # empty
                'Normal message',                   # neutral
                '<script>alert("test")</script>',   # HTML content
            ],
            'sender': ['u1'] * 5,
            'timestamp': pd.date_range(start='2024-01-01', periods=5, freq='h'),
        })

        results = analyzer.analyze_sentiment(test_data)
        assert len(results) == 5, "Should handle all edge cases without crashing"
        assert 'sentiment_polarity' in results.columns