#!/usr/bin/env python3
"""Test Excel filtering to only show mapped persons."""

import json
import glob
from pathlib import Path
import pandas as pd
from src.reporters.excel_reporter import ExcelReporter
from src.forensic_utils import ForensicRecorder
from src.config import Config

# Load config
config = Config()
print(f"Mapped persons: {list(config.contact_mappings.keys())}")

# Initialize recorder
recorder = ForensicRecorder(Path(config.output_dir))

# Load latest extracted data and analysis
latest_data = sorted(glob.glob(str(Path(config.output_dir) / 'extracted_data_*.json')))[-1]
latest_analysis = sorted(glob.glob(str(Path(config.output_dir) / 'analysis_results_*.json')))[-1]

print(f"Loading data from: {latest_data}")
print(f"Loading analysis from: {latest_analysis}")

with open(latest_data) as f:
    data = json.load(f)
with open(latest_analysis) as f:
    analysis = json.load(f)

print(f"Total messages in data: {len(data.get('messages', []))}")

# Check recipient distribution
df = pd.DataFrame(data['messages'])
print(f"\nUnique recipients: {df['recipient'].nunique()}")
print(f"Top 10 recipients by message count:")
print(df['recipient'].value_counts().head(10))

# Enrich with analysis data
df_messages = df.copy()
if 'threats' in analysis and 'details' in analysis['threats']:
    df_threats = pd.DataFrame(analysis['threats']['details'])
    if 'message_id' in df_messages.columns and 'message_id' in df_threats.columns:
        threat_cols = [col for col in df_threats.columns if col.startswith('threat_') or col == 'harmful_content']
        if 'message_id' not in threat_cols:
            threat_cols.insert(0, 'message_id')
        df_messages = df_messages.merge(df_threats[threat_cols], on='message_id', how='left')

if 'sentiment' in analysis:
    df_sentiment = pd.DataFrame(analysis['sentiment'])
    if 'message_id' in df_sentiment.columns:
        sentiment_cols = [col for col in df_sentiment.columns if col.startswith('sentiment_')]
        if 'message_id' not in sentiment_cols:
            sentiment_cols.insert(0, 'message_id')
        df_messages = df_messages.merge(df_sentiment[sentiment_cols], on='message_id', how='left')

enriched_data = data.copy()
enriched_data['messages'] = df_messages.to_dict('records')

# Generate test Excel
excel_reporter = ExcelReporter(recorder)
test_path = Path(config.output_dir) / 'test_filtered_excel.xlsx'
print(f"\nGenerating Excel to: {test_path}")

excel_reporter.generate_report(enriched_data, analysis, {}, test_path)

# Check results
import openpyxl
wb = openpyxl.load_workbook(test_path)
print(f"\nGenerated Excel with sheets: {wb.sheetnames}")
print(f"Number of sheets: {len(wb.sheetnames)}")
print("\nSheet row counts:")
for name in wb.sheetnames:
    print(f"  {name}: {wb[name].max_row} rows")
