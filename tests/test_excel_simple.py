#!/usr/bin/env python3
"""Simple test of Excel filtering logic."""

from pathlib import Path
import pandas as pd
from src.reporters.excel_reporter import ExcelReporter
from src.forensic_utils import ForensicRecorder
from src.config import Config

# Load config
config = Config()
print(f"Mapped persons: {list(config.contact_mappings.keys())}")

# Create test data with mapped and unmapped persons
test_messages = [
    {'message_id': 1, 'sender': 'Me', 'recipient': 'Marcia Snyder', 'content': 'Test to Marcia', 'timestamp': '2024-01-01', 'source': 'imessage'},
    {'message_id': 2, 'sender': 'Marcia Snyder', 'recipient': 'Me', 'content': 'Reply from Marcia', 'timestamp': '2024-01-01', 'source': 'imessage'},
    {'message_id': 3, 'sender': 'Me', 'recipient': 'Kiara Snyder', 'content': 'Test to Kiara', 'timestamp': '2024-01-01', 'source': 'imessage'},
    {'message_id': 4, 'sender': 'Kiara Snyder', 'recipient': 'Me', 'content': 'Reply from Kiara', 'timestamp': '2024-01-01', 'source': 'imessage'},
    {'message_id': 5, 'sender': 'Me', 'recipient': '+12065551234', 'content': 'Test to random number', 'timestamp': '2024-01-01', 'source': 'imessage'},
    {'message_id': 6, 'sender': '+12065551234', 'recipient': 'Me', 'content': 'Reply from random', 'timestamp': '2024-01-01', 'source': 'imessage'},
    {'message_id': 7, 'sender': 'Me', 'recipient': 'chat123456', 'content': 'Test to chat', 'timestamp': '2024-01-01', 'source': 'imessage'},
]

print(f"\nTest data has {len(test_messages)} messages")
df = pd.DataFrame(test_messages)
print(f"Unique recipients: {sorted(df['recipient'].unique())}")

# Create test data structure
test_data = {
    'messages': test_messages,
    'total_messages': len(test_messages),
    'screenshots': []
}

test_analysis = {
    'threats': {'details': [], 'summary': {}},
    'sentiment': []
}

# Initialize recorder
recorder = ForensicRecorder(Path(config.output_dir))

# Generate Excel
excel_reporter = ExcelReporter(recorder)
test_path = Path(config.output_dir) / 'test_simple_excel.xlsx'
print(f"\nGenerating Excel to: {test_path}")

excel_reporter.generate_report(test_data, test_analysis, {}, test_path)

# Check results
import openpyxl
wb = openpyxl.load_workbook(test_path)
print(f"\n✓ Generated Excel successfully")
print(f"Sheet names: {wb.sheetnames}")
print(f"Number of sheets: {len(wb.sheetnames)}")

print("\nExpected: Overview, Marcia Snyder, Kiara Snyder, All Messages")
print(f"Expected NOT to see: +12065551234, chat123456")

if '+12065551234' in wb.sheetnames or 'chat123456' in wb.sheetnames:
    print("\n❌ FAIL: Excel contains unmapped recipients!")
else:
    print("\n✓ PASS: Excel only contains mapped persons!")

print("\nSheet details:")
for name in wb.sheetnames:
    row_count = wb[name].max_row
    print(f"  {name}: {row_count} rows")
